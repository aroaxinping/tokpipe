"""Generate Excel reports with native formulas and formatting."""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from tokpipe.metrics import Report


HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="ffffff", size=11)
DATA_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def to_excel(
    report: Report,
    path: str | Path,
    followers: int | None = None,
    period: str | None = None,
) -> None:
    """Export report to Excel with formulas, formatting, and charts.

    Args:
        report: Computed metrics report.
        path: Output path for the .xlsx file.
        followers: Total follower count (used for reach rate formulas).
        period: Analysis period label (e.g., "24 Feb - 23 Mar 2026").
    """
    wb = Workbook()

    _write_summary_sheet(wb, report, followers, period)
    _write_data_sheet(wb, report)
    _write_charts_sheet(wb, report)

    wb.save(path)


def _style_header_row(ws, num_cols: int) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_summary_sheet(wb, report: Report, followers: int | None, period: str | None) -> None:
    ws = wb.active
    ws.title = "Summary"

    rows = [
        ("Metric", "Value"),
        ("Total videos", len(report.data)),
        ("Avg engagement rate", f"=Data!{get_column_letter(report.data.shape[1] + 1)}2"),
    ]

    if period:
        rows.insert(1, ("Period", period))

    if followers:
        rows.append(("Followers", followers))

    if report.best_hour is not None:
        rows.append(("Best posting hour", f"{report.best_hour}:00"))

    rows.append(("Top performers", len(report.top_performers)))

    for r, (label, value) in enumerate(rows, start=1):
        ws.cell(row=r, column=1, value=label).font = DATA_FONT
        ws.cell(row=r, column=2, value=value).font = DATA_FONT

    _style_header_row(ws, 2)
    _auto_width(ws)


def _write_data_sheet(wb, report: Report) -> None:
    ws = wb.create_sheet("Data")

    df = report.data.copy()

    # Write headers
    headers = list(df.columns) + ["engagement_rate"]
    if report.completion_rate is not None:
        headers.append("completion_rate")

    for c, header in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=header)

    _style_header_row(ws, len(headers))

    # Write data rows
    for r, (_, row) in enumerate(df.iterrows(), start=2):
        for c, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                ws.cell(row=r, column=c, value="")
            elif hasattr(val, "isoformat"):
                ws.cell(row=r, column=c, value=val)
            else:
                ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).font = DATA_FONT
            ws.cell(row=r, column=c).border = THIN_BORDER

    # Engagement rate column with formulas
    er_col = len(df.columns) + 1
    er_letter = get_column_letter(er_col)

    # Find column letters for views, likes, comments, shares
    from tokpipe.metrics import _find_column
    views_col = _find_column(df, ["views", "view", "reproduc"])
    likes_col = _find_column(df, ["likes", "like", "me_gusta"])
    comments_col = _find_column(df, ["comments", "comment", "comentario"])
    shares_col = _find_column(df, ["shares", "share", "compartid"])

    col_indices = {col: i + 1 for i, col in enumerate(df.columns)}

    if views_col:
        v_let = get_column_letter(col_indices[views_col])
        parts = []
        if likes_col:
            parts.append(get_column_letter(col_indices[likes_col]))
        if comments_col:
            parts.append(get_column_letter(col_indices[comments_col]))
        if shares_col:
            parts.append(get_column_letter(col_indices[shares_col]))

        for r in range(2, len(df) + 2):
            numerator = "+".join(f"{p}{r}" for p in parts) if parts else "0"
            formula = f'=IF({v_let}{r}=0,0,({numerator})/{v_let}{r})'
            ws.cell(row=r, column=er_col, value=formula)
            ws.cell(row=r, column=er_col).font = DATA_FONT
            ws.cell(row=r, column=er_col).border = THIN_BORDER
            ws.cell(row=r, column=er_col).number_format = "0.00%"

    # Average formula at top of summary
    avg_formula = f"=AVERAGE({er_letter}2:{er_letter}{len(df) + 1})"
    summary = wb["Summary"]
    for row in summary.iter_rows(min_row=1, max_row=summary.max_row, max_col=2):
        if row[0].value == "Avg engagement rate":
            row[1].value = avg_formula
            row[1].number_format = "0.00%"

    _auto_width(ws)


def _write_charts_sheet(wb, report: Report) -> None:
    ws = wb.create_sheet("Charts")
    data_ws = wb["Data"]
    num_rows = len(report.data) + 1

    # Find engagement_rate column
    er_col = data_ws.max_column

    # Engagement rate bar chart
    chart = BarChart()
    chart.title = "Engagement Rate per Video"
    chart.y_axis.title = "Engagement Rate"
    chart.x_axis.title = "Video"
    chart.style = 10
    chart.width = 25
    chart.height = 12

    data_ref = Reference(data_ws, min_col=er_col, min_row=1, max_row=num_rows)
    chart.add_data(data_ref, titles_from_data=True)
    ws.add_chart(chart, "A1")
