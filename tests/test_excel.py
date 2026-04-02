"""Tests for tokpipe.excel module."""

import pandas as pd
import pytest
from openpyxl import load_workbook

from tokpipe.excel import to_excel
from tokpipe.metrics import Report


def _make_report(n_rows=5):
    """Build a simple Report object for testing."""
    df = pd.DataFrame({
        "views": [1000, 2000, 500, 800, 1500][:n_rows],
        "likes": [100, 300, 50, 80, 200][:n_rows],
        "comments": [10, 30, 5, 8, 20][:n_rows],
        "shares": [5, 15, 2, 4, 10][:n_rows],
    })
    views = df["views"]
    safe_views = views.replace(0, pd.NA).astype(float)
    engagement_rate = (df["likes"] + df["comments"] + df["shares"]) / safe_views
    engagement_rate = engagement_rate.fillna(0)
    like_rate = (df["likes"] / safe_views).fillna(0)
    comment_rate = (df["comments"] / safe_views).fillna(0)
    share_rate = (df["shares"] / safe_views).fillna(0)
    threshold = engagement_rate.quantile(0.9)
    top_performers = df[engagement_rate >= threshold].copy()

    return Report(
        data=df,
        engagement_rate=engagement_rate,
        like_rate=like_rate,
        comment_rate=comment_rate,
        share_rate=share_rate,
        save_rate=None,
        virality_score=share_rate.copy(),
        views_per_day=None,
        avg_watch_time=None,
        completion_rate=None,
        best_hour=14,
        growth_trend=None,
        top_performers=top_performers,
    )


class TestToExcelGeneratesFile:
    def test_creates_valid_xlsx(self, tmp_path):
        path = tmp_path / "report.xlsx"
        to_excel(_make_report(), path)
        assert path.exists()
        assert path.stat().st_size > 0
        # Verify openpyxl can open it without error
        wb = load_workbook(path)
        wb.close()


class TestSheetNames:
    def test_has_expected_sheets(self, tmp_path):
        path = tmp_path / "report.xlsx"
        to_excel(_make_report(), path)
        wb = load_workbook(path)
        assert wb.sheetnames == ["Summary", "Data", "Charts"]
        wb.close()


class TestSummarySheet:
    def test_contains_total_videos(self, tmp_path):
        path = tmp_path / "report.xlsx"
        report = _make_report()
        to_excel(report, path)
        wb = load_workbook(path)
        ws = wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Total videos" in labels
        row_idx = labels.index("Total videos") + 1
        assert ws.cell(row=row_idx, column=2).value == len(report.data)
        wb.close()

    def test_contains_best_posting_hour(self, tmp_path):
        path = tmp_path / "report.xlsx"
        report = _make_report()
        to_excel(report, path)
        wb = load_workbook(path)
        ws = wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Best posting hour" in labels
        row_idx = labels.index("Best posting hour") + 1
        assert ws.cell(row=row_idx, column=2).value == "14:00"
        wb.close()

    def test_contains_top_performers(self, tmp_path):
        path = tmp_path / "report.xlsx"
        report = _make_report()
        to_excel(report, path)
        wb = load_workbook(path)
        ws = wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Top performers" in labels
        wb.close()

    def test_contains_avg_engagement_rate_formula(self, tmp_path):
        path = tmp_path / "report.xlsx"
        to_excel(_make_report(), path)
        wb = load_workbook(path)
        ws = wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        row_idx = labels.index("Avg engagement rate") + 1
        value = ws.cell(row=row_idx, column=2).value
        assert isinstance(value, str) and value.startswith("=")
        wb.close()


class TestDataSheet:
    def test_correct_number_of_rows(self, tmp_path):
        path = tmp_path / "report.xlsx"
        report = _make_report(n_rows=3)
        to_excel(report, path)
        wb = load_workbook(path)
        ws = wb["Data"]
        # 1 header row + 3 data rows
        data_rows = ws.max_row - 1
        assert data_rows == 3
        wb.close()

    def test_engagement_rate_column_has_formulas(self, tmp_path):
        path = tmp_path / "report.xlsx"
        to_excel(_make_report(), path)
        wb = load_workbook(path)
        ws = wb["Data"]
        # Find the engagement_rate column
        er_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "engagement_rate":
                er_col = col
                break
        assert er_col is not None, "engagement_rate column not found"
        # Every data row should contain a formula
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=er_col).value
            assert isinstance(val, str) and val.startswith("="), (
                f"Row {row} engagement_rate is not a formula: {val!r}"
            )
        wb.close()

    def test_headers_include_original_columns_and_engagement_rate(self, tmp_path):
        path = tmp_path / "report.xlsx"
        to_excel(_make_report(), path)
        wb = load_workbook(path)
        ws = wb["Data"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert "views" in headers
        assert "likes" in headers
        assert "engagement_rate" in headers
        wb.close()


class TestOptionalParams:
    def test_with_followers_and_period(self, tmp_path):
        path = tmp_path / "report.xlsx"
        to_excel(_make_report(), path, followers=12000, period="24 Feb - 23 Mar 2026")
        wb = load_workbook(path)
        ws = wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        values = {
            ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
            for r in range(1, ws.max_row + 1)
        }
        assert "Followers" in labels
        assert values["Followers"] == 12000
        assert "Period" in labels
        assert values["Period"] == "24 Feb - 23 Mar 2026"
        wb.close()

    def test_without_optional_params(self, tmp_path):
        path = tmp_path / "report.xlsx"
        to_excel(_make_report(), path, followers=None, period=None)
        wb = load_workbook(path)
        ws = wb["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Followers" not in labels
        assert "Period" not in labels
        wb.close()
